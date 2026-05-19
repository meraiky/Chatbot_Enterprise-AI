"""
data_ingestion.py — Document ingestion with MinerU-first extraction.

Extraction priority:
  1. MinerU (magic-pdf) — layout-aware, handles tables/formulas/images,
     outputs clean Markdown.  Requires ``pip install magic-pdf[full]``.
  2. PyMuPDF (fitz) — legacy fallback when MinerU is not installed or fails.

The rest of the pipeline (chunking → injection scan → pgvector indexing →
usage tracking → cache invalidation) is shared.
"""

from __future__ import annotations

import json
import logging
import re
import csv
import zipfile
import tempfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import List

import fitz  # PyMuPDF
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_core.documents import Document
from app.services.rag.document_registry import (
    build_document_id,
    delete_existing_document,
    utc_now,
)
from app.core.config import settings
from app.services.usage_tracker import estimate_tokens, new_request_id, record_usage
from app.services.rag.vector_store import get_vector_store
from app.services.rag.cache_service import cache_service
from app.services.rag.document_mirror import mirror_document_chunks
from app.services.rag.injection_scanner import sanitize_chunk, scan_chunk

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Optional MinerU import — graceful degradation to PyMuPDF
# ---------------------------------------------------------------------------
_MINERU_AVAILABLE = False
try:
    from magic_pdf.data.data_reader_writer import FileBasedDataWriter, FileBasedDataReader
    from magic_pdf.data.dataset import PymuDocDataset
    from magic_pdf.model.doc_analyze_by_custom_model import doc_analyze
    _MINERU_AVAILABLE = True
    logger.info("MinerU (magic-pdf) loaded — using advanced document extraction")
except ImportError:
    logger.info("MinerU not installed — falling back to PyMuPDF extraction")

INDEX_BATCH_SIZE = 20
SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".csv", ".xlsx"}

# Supported file extensions for MinerU multi-format ingestion
MINERU_SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".pptx", ".xlsx", ".doc"}


# ---------------------------------------------------------------------------
# Extraction backends
# ---------------------------------------------------------------------------

def _extract_with_mineru(file_path: str, doc_id: str, doc_type: str,
                         source_name: str, checksum: str,
                         uploaded_at: str) -> List[Document]:
    """Extract document content using MinerU for layout-aware Markdown output.

    MinerU performs:
      - Layout analysis (headers, paragraphs, lists, tables, images)
      - OCR on scanned pages
      - Table structure recognition
      - Formula detection
      - Clean Markdown output optimised for LLM consumption
    """
    file_ext = Path(file_path).suffix.lower()
    if file_ext not in MINERU_SUPPORTED_EXTENSIONS:
        raise ValueError(f"MinerU does not support '{file_ext}' files")

    with tempfile.TemporaryDirectory(prefix="mineru_") as tmp_dir:
        tmp_path = Path(tmp_dir)
        output_dir = tmp_path / "output"
        output_dir.mkdir()

        # Read the raw file bytes
        reader = FileBasedDataReader("")
        file_bytes = Path(file_path).read_bytes()

        # Run MinerU analysis pipeline
        ds = PymuDocDataset(file_bytes)
        infer_result = ds.apply(doc_analyze, ocr=True)

        # Write Markdown output
        writer = FileBasedDataWriter(str(output_dir))
        md_content = infer_result.get_markdown(writer)

    if not md_content or not md_content.strip():
        raise ValueError("MinerU extracted no content from the document.")

    # Split the Markdown into page-like sections using heading boundaries
    # MinerU output uses ## for page/section breaks
    sections = re.split(r'\n(?=#{1,3}\s)', md_content)
    pages: List[Document] = []
    for page_number, section in enumerate(sections, start=1):
        section_text = section.strip()
        if section_text:
            pages.append(
                Document(
                    page_content=section_text,
                    metadata={
                        "doc_id": doc_id,
                        "source": source_name,
                        "type": doc_type,
                        "page": page_number,
                        "checksum": checksum,
                        "uploaded_at": uploaded_at,
                        "extractor": "mineru",
                    },
                )
            )
    return pages


def _extract_with_pymupdf(file_path: str, doc_id: str, doc_type: str,
                           source_name: str, checksum: str,
                           uploaded_at: str) -> List[Document]:
    """Legacy PyMuPDF extraction — plain text, no layout awareness."""
    pages: List[Document] = []
    with fitz.open(file_path) as doc:
        for page_number, page in enumerate(doc, start=1):
            page_text = page.get_text().strip()
            page_text = re.sub(r'\s+', ' ', page_text).strip()
            if page_text:
                pages.append(
                    Document(
                        page_content=page_text,
                        metadata={
                            "doc_id": doc_id,
                            "source": source_name,
                            "type": doc_type,
                            "page": page_number,
                            "checksum": checksum,
                            "uploaded_at": uploaded_at,
                            "extractor": "pymupdf",
                        },
                    )
                )
    return pages


def _docx_documents(file_path: str, doc_id: str, doc_type: str, source_name: str, checksum: str, uploaded_at: str) -> list[Document]:
    try:
        with zipfile.ZipFile(file_path) as archive:
            xml_bytes = archive.read("word/document.xml")
    except Exception as exc:
        raise ValueError(f"Failed to read DOCX: {exc}") from exc

    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    root = ET.fromstring(xml_bytes)
    blocks: list[str] = []

    for paragraph in root.findall(".//w:p", namespace):
        text = "".join(node.text or "" for node in paragraph.findall(".//w:t", namespace)).strip()
        if text:
            blocks.append(text)

    text = "\n".join(blocks).strip()
    if not text:
        return []

    return [
        Document(
            page_content=text,
            metadata={
                "doc_id": doc_id,
                "source": source_name,
                "type": doc_type,
                "page": 1,
                "checksum": checksum,
                "uploaded_at": uploaded_at,
                "file_type": "docx",
                "extractor": "manual",
            },
        )
    ]


def _csv_documents(file_path: str, doc_id: str, doc_type: str, source_name: str, checksum: str, uploaded_at: str) -> list[Document]:
    encodings = ("utf-8-sig", "utf-8", "cp1258", "latin-1")
    rows: list[list[str]] | None = None
    last_error: Exception | None = None
    for encoding in encodings:
        try:
            with open(file_path, newline="", encoding=encoding) as handle:
                rows = list(csv.reader(handle))
            break
        except Exception as exc:
            last_error = exc

    if rows is None:
        raise ValueError(f"Failed to read CSV: {last_error}")

    lines = [" | ".join(cell.strip() for cell in row if cell is not None).strip() for row in rows]
    text = "\n".join(line for line in lines if line).strip()
    if not text:
        return []

    return [
        Document(
            page_content=text,
            metadata={
                "doc_id": doc_id,
                "source": source_name,
                "type": doc_type,
                "page": 1,
                "checksum": checksum,
                "uploaded_at": uploaded_at,
                "file_type": "csv",
                "extractor": "manual",
            },
        )
    ]


def _xlsx_documents(file_path: str, doc_id: str, doc_type: str, source_name: str, checksum: str, uploaded_at: str) -> list[Document]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("openpyxl is not installed. Install backend requirements to upload Excel files.") from exc

    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Failed to read XLSX: {exc}") from exc

    documents: list[Document] = []
    try:
        for sheet_index, sheet in enumerate(workbook.worksheets, start=1):
            lines: list[str] = [f"Sheet: {sheet.title}"]
            for row in sheet.iter_rows(values_only=True):
                values = [str(value).strip() for value in row if value is not None and str(value).strip()]
                if values:
                    lines.append(" | ".join(values))
            text = "\n".join(lines).strip()
            if len(lines) > 1:
                documents.append(
                    Document(
                        page_content=text,
                        metadata={
                            "doc_id": doc_id,
                            "source": source_name,
                            "type": doc_type,
                            "page": sheet_index,
                            "sheet": sheet.title,
                            "checksum": checksum,
                            "uploaded_at": uploaded_at,
                            "file_type": "xlsx",
                            "extractor": "manual",
                        },
                    )
                )
    finally:
        workbook.close()

    return documents


# ---------------------------------------------------------------------------
# Main ingestion entry point
# ---------------------------------------------------------------------------

def process_and_index_file(file_path: str, doc_type: str, source_name: str, checksum: str):
    """
    Ingest a document (PDF/DOCX/CSV/XLSX) into the vector store.

    Extraction priority for PDFs:
      1. MinerU — if installed and file format is supported
      2. PyMuPDF — fallback for PDFs when MinerU is unavailable

    For DOCX/CSV/XLSX: uses manual extractors (unchanged).

    The function handles chunking, injection scanning, pgvector indexing,
    usage tracking, and cache invalidation.
    """
    doc_id = build_document_id(source_name, doc_type, checksum)
    uploaded_at = utc_now()
    extractor_used = "unknown"

    # ── Step 1: Extract pages ──────────────────────────────────────────────
    pages: List[Document] = []
    file_ext = Path(source_name).suffix.lower()

    # Try MinerU first for supported formats
    if _MINERU_AVAILABLE and file_ext in MINERU_SUPPORTED_EXTENSIONS:
        try:
            pages = _extract_with_mineru(
                file_path, doc_id, doc_type, source_name, checksum, uploaded_at
            )
            extractor_used = "mineru"
            logger.info("MinerU extraction succeeded for %s (%d sections)",
                        source_name, len(pages))
        except Exception as e:
            logger.warning("MinerU extraction failed for %s, falling back: %s",
                           source_name, e)
            pages = []

    # Fallback to format-specific extractors
    if not pages:
        if file_ext == ".pdf":
            try:
                pages = _extract_with_pymupdf(
                    file_path, doc_id, doc_type, source_name, checksum, uploaded_at
                )
                extractor_used = "pymupdf"
                logger.info("PyMuPDF extraction succeeded for %s (%d pages)",
                            source_name, len(pages))
            except Exception as e:
                raise ValueError(f"Failed to read PDF: {e}")
        elif file_ext == ".docx":
            pages = _docx_documents(file_path, doc_id, doc_type, source_name, checksum, uploaded_at)
            extractor_used = "manual"
        elif file_ext == ".csv":
            pages = _csv_documents(file_path, doc_id, doc_type, source_name, checksum, uploaded_at)
            extractor_used = "manual"
        elif file_ext == ".xlsx":
            pages = _xlsx_documents(file_path, doc_id, doc_type, source_name, checksum, uploaded_at)
            extractor_used = "manual"
        elif file_ext == ".doc":
            raise ValueError("Legacy .doc files are not supported. Please save the file as .docx and upload again.")
        elif file_ext == ".xls":
            raise ValueError("Legacy .xls files are not supported. Please save the file as .xlsx or .csv and upload again.")
        else:
            raise ValueError(f"Unsupported file type: {file_ext or 'unknown'}")

    if not pages:
        raise ValueError("Document does not contain extractable text.")

    # ── Step 2: Chunk ──────────────────────────────────────────────────────
    text_splitter = RecursiveCharacterTextSplitter(
        chunk_size=1000,
        chunk_overlap=100,
        length_function=len
    )

    documents = text_splitter.split_documents(pages)
    documents = [doc for doc in documents if doc.page_content.strip()]

    # ── Step 3: Injection scan ─────────────────────────────────────────────
    for document in documents:
        scan = scan_chunk(document.page_content)
        if scan["risk_score"] > 0:
            document.page_content = sanitize_chunk(document.page_content)
            document.metadata["injection_sanitized"] = True
            document.metadata["injection_findings"] = ",".join(scan["findings"])

    if not documents:
        raise ValueError("PDF text was extracted, but no indexable chunks were created.")

    # ── Step 4: Index in pgvector ──────────────────────────────────────────
    vector_store = get_vector_store()
    replaced_chunks = delete_existing_document(doc_id)
    all_ids = []
    for start in range(0, len(documents), INDEX_BATCH_SIZE):
        batch = documents[start : start + INDEX_BATCH_SIZE]
        ids = [
            f"{doc_id}:chunk-{start + index + 1}"
            for index in range(len(batch))
        ]
        all_ids.extend(ids)
        vector_store.add_documents(batch, ids=ids)

    # ── Step 5: Mirror to document store ───────────────────────────────────
    mirror_document_chunks(documents=documents, ids=all_ids)

    # ── Step 6: Usage tracking ─────────────────────────────────────────────
    request_id = new_request_id()
    embedding_tokens = sum(estimate_tokens(document.page_content) for document in documents)
    usage = record_usage(
        request_id=request_id,
        operation="document_embedding",
        mode=doc_type,
        model=settings.EMBEDDING_MODEL,
        input_tokens=embedding_tokens,
        estimated=True,
        metadata=json.dumps({
            "doc_id": doc_id,
            "source": source_name,
            "extractor": extractor_used,
        }),
    )
    
    # ── Step 7: Cache invalidation ─────────────────────────────────────────
    cache_service.clear_cache_by_mode(doc_type)
    
    return {
        "doc_id": doc_id,
        "chunks_indexed": len(documents),
        "replaced_chunks": replaced_chunks,
        "extractor": extractor_used,
        "usage": usage,
        "uploaded_at": uploaded_at,
    }


# Backward-compatible alias so existing callers keep working
def process_and_index_pdf(file_path: str, doc_type: str, source_name: str, checksum: str):
    return process_and_index_file(file_path, doc_type, source_name, checksum)
